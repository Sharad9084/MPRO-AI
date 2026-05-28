from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def export_rows_to_excel(rows: list[dict[str, Any]], output_path: Path, sheet_name: str = "Extracted Data") -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name[:31] or "Extracted Data"

    columns = _columns_from_rows(rows)
    sheet.append(columns)
    for row in rows:
        sheet.append([row.get(column, "") for column in columns])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    sheet.freeze_panes = "A2"
    for index, column in enumerate(columns, start=1):
        width = max(len(str(column)), *(len(str(row.get(column, ""))) for row in rows), 12)
        sheet.column_dimensions[get_column_letter(index)].width = min(width + 2, 45)

    workbook.save(output_path)
    return output_path


def _columns_from_rows(rows: list[dict[str, Any]]) -> list[str]:
    preferred = [
        "Source Type",
        "PDF File Name",
        "Row No",
        "Advertiser Name",
        "Agency Name",
        "Medium",
        "Campaign Period",
        "PO Number",
        "PO Date",
        "Invoice Number",
        "Invoice Date",
        "Brand",
        "Brand Name",
        "Description",
        "Campaign Name",
        "PO Amount Incl Tax",
        "Total Value Including Taxes",
        "Channel Name",
        "Program",
        "Date",
        "Air Time",
        "Spot Duration",
        "Rate INR",
        "Final Amount INR",
        "Parser Confidence",
        "Extraction Status",
    ]
    seen = set()
    columns: list[str] = []
    for column in preferred:
        if any(column in row for row in rows):
            columns.append(column)
            seen.add(column)
    for row in rows:
        for column in row:
            if column not in seen:
                columns.append(column)
                seen.add(column)
    return columns

